import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

from datetime import datetime
import time
import numpy as np
import scipy.sparse as sp
from sklearn.cluster import KMeans
from skimage import io, color
import os
from sklearn.neighbors import NearestNeighbors
import pandas as pd
import re
from openpyxl import load_workbook
from scipy.sparse.linalg import eigsh

def compute_weight_matrix_coo_knn(image, sigma_i, sigma_x, k_neighbors=10):
    h, w, c = image.shape
    N = h * w

    coords = np.array(np.meshgrid(range(h), range(w))).reshape(2, -1).T
    features = image.reshape(-1, c)

    knn = NearestNeighbors(n_neighbors=k_neighbors, algorithm="ball_tree")
    knn.fit(coords)
    distances, indices = knn.kneighbors(coords)

    row_idx = np.repeat(np.arange(N), k_neighbors)
    col_idx = indices.flatten()

    feat_row = features[row_idx]
    feat_col = features[col_idx]

    diff = feat_row - feat_col
    W_feature = np.exp(-np.sum(diff ** 2, axis=1) / (2 * sigma_i ** 2))
    W_space = np.exp(-(distances.flatten() ** 2) / (2 * sigma_x ** 2))

    values = W_feature * W_space

    W_sparse = sp.coo_matrix((values, (row_idx, col_idx)), shape=(N, N))
    return W_sparse


def compute_laplacian_coo(W_coo):
    D = np.array(W_coo.sum(axis=1)).flatten()
    D_coo = sp.diags(D)
    L_coo = D_coo - W_coo
    return L_coo, D_coo

def compute_ncut_lanczos_cpu(W_coo, k=2, max_iter=100, tol=1e-5):
    """
    Tính k vector riêng nhỏ nhất của A = I - D^{-1/2} W D^{-1/2}
    bằng phương pháp Lanczos, chạy trên CPU với NumPy và SciPy.
    """

    n = W_coo.shape[0]

    # 1. Chuẩn hóa W: W_norm = D^{-1/2} * W * D^{-1/2}
    D_vals = np.array(W_coo.sum(axis=1)).flatten()
    D_inv_sqrt = 1.0 / np.sqrt(D_vals + 1e-8)
    row, col = W_coo.row, W_coo.col
    data = W_coo.data * D_inv_sqrt[row] * D_inv_sqrt[col]
    W_norm = coo_matrix((data, (row, col)), shape=W_coo.shape)

    # 2. Định nghĩa hàm nhân A = I - W_norm
    def A_mul(x):
        return x - W_norm @ x

    # 3. Khởi tạo vector đầu
    Q = []
    alphas = []
    betas = []

    q = np.random.randn(n).astype(np.float32)
    q /= np.linalg.norm(q)
    Q.append(q)
    beta = 0.0
    q_prev = np.zeros_like(q)

    for j in range(max_iter):
        z = A_mul(Q[-1])
        alpha = np.dot(Q[-1], z)
        alphas.append(alpha)

        z = z - alpha * Q[-1] - beta * q_prev

        # Re-orthogonalization
        for q_i in Q:
            z -= np.dot(q_i, z) * q_i

        beta = np.linalg.norm(z)
        if beta < tol or len(alphas) >= k + 50:
            break

        betas.append(beta)
        q_prev = Q[-1]
        Q.append(z / beta)

    # 4. Ma trận tridiagonal T
    m = len(alphas)
    T = np.zeros((m, m), dtype=np.float32)
    for i in range(m):
        T[i, i] = alphas[i]
        if i > 0:
            T[i, i-1] = T[i-1, i] = betas[i-1]

    # 5. Giải trị riêng
    vals, vecs = np.linalg.eigh(T)

    sorted_idx = np.argsort(vals)
    vecs = vecs[:, sorted_idx]

    nonzero = np.where(np.abs(vals[sorted_idx]) > 1e-5)[0]
    vecs = vecs[:, nonzero[:k]]

    # 6. Trả về vector riêng trong không gian gốc
    Q_mat = np.stack(Q, axis=1)  # (n, m)
    eigenvectors = Q_mat @ vecs  # (n, k)

    return eigenvectors

def assign_labels(eigen_vectors, k):
    return KMeans(n_clusters=k, random_state=0).fit(eigen_vectors).labels_

def save_segmentation(image, labels, k, output_path):
    h, w, c = image.shape
    segmented_image = np.zeros_like(image, dtype=np.uint8)
    for i in range(k):
        mask = labels.reshape(h, w) == i
        cluster_pixels = image[mask]
        mean_color = (cluster_pixels.mean(axis=0) * 255).astype(np.uint8) if len(cluster_pixels) > 0 else np.array([0, 0, 0], dtype=np.uint8)
        segmented_image[mask] = mean_color
    io.imsave(output_path, segmented_image)

def save_seg_file(labels, image_shape, output_path, image_name="image"):
    h, w = image_shape[:2]
    unique_labels = np.unique(labels)
    segments = len(unique_labels)

    # Tạo phần header
    header = [
        "format ascii cr",
        f"date {datetime.now().strftime('%a %b %d %H:%M:%S %Y')}",
        f"image ",
        "user 1102",  # Giữ nguyên như file mẫu
        f"width {w}",
        f"height {h}",
        f"segments {segments}",
        "gray 0",
        "invert 0",
        "flipflop 0",
        "data"
    ]

    # Tạo dữ liệu pixel theo định dạng (nhãn, dòng, cột bắt đầu, cột kết thúc)
    data_lines = []
    for row in range(h):
        row_labels = labels[row, :]
        start_col = 0
        current_label = row_labels[0]

        for col in range(1, w):
            if row_labels[col] != current_label:
                data_lines.append(f"{current_label} {row} {start_col} {col}")
                start_col = col
                current_label = row_labels[col]

        # Thêm dòng cuối cùng của hàng
        data_lines.append(f"{current_label} {row} {start_col} {w}")

    # Lưu vào file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(header) + "\n")
        f.write("\n".join(data_lines) + "\n")

    print(f"✅ File SEG đã lưu: {output_path}")

def align_eigenvector_signs(vecs, V_ql, V_qpe, V_iqpe):
    """
    Align signs of V_ql, V_qpe, V_iqpe to match reference eigenvectors vecs.

    Parameters
    ----------
    vecs : np.ndarray (N,k)
        Reference eigenvectors (traditional Lanczos).
    V_ql : np.ndarray (N,k)
    V_qpe : np.ndarray (N,k)
    V_iqpe : np.ndarray (N,k)

    Returns
    -------
    V_ql_aligned, V_qpe_aligned, V_iqpe_aligned
    """

    k = vecs.shape[1]

    V_ql_aligned = V_ql.copy()
    V_qpe_aligned = V_qpe.copy()
    V_iqpe_aligned = V_iqpe.copy()

    for i in range(k):
        ref = vecs[:, i]

        if np.dot(ref, V_ql_aligned[:, i]) < 0:
            V_ql_aligned[:, i] *= -1

        if np.dot(ref, V_qpe_aligned[:, i]) < 0:
            V_qpe_aligned[:, i] *= -1

        if np.dot(ref, V_iqpe_aligned[:, i]) < 0:
            V_iqpe_aligned[:, i] *= -1

    return V_ql_aligned, V_qpe_aligned, V_iqpe_aligned

def normalized_cuts_eigsh(imagename, image_path, output_path, k, sigma_i, sigma_x):
    image = io.imread(image_path)
    image = color.gray2rgb(image) if image.ndim == 2 else image[:, :, :3] if image.shape[2] == 4 else image
    image = image / 255.0

    start_vecs = time.perf_counter()
    W_coo = compute_weight_matrix_coo_knn(image, sigma_i, sigma_x)
    _ = compute_ncut_lanczos_cpu(W_coo, k)
    end_vecs = time.perf_counter()

    evals, vecs= smallest_eigenpairs_ncut(W_coo, k)

    E_ql, E_qpe, E_iqpe, V_ql, V_qpe, V_iqpe, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe = compute_ncut_ql_pipeline_three(W_coo, k)

    V_ql, V_qpe, V_iqpe = align_eigenvector_signs(vecs, V_ql, V_qpe, V_iqpe)

    labels = assign_labels(vecs, k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "____L.seg", imagename)

    labels = assign_labels(V_ql, k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "___QL.seg", imagename)

    labels = assign_labels(V_qpe, k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "__QPE.seg", imagename)

    labels = assign_labels(V_iqpe, k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_IQPE.seg", imagename)
    
    del W_coo

    return start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe, evals, E_ql, E_qpe, E_iqpe, vecs, V_ql, V_qpe, V_iqpe

def append_eigenvectors_row_format(
    excel_path,
    file_name,
    vecs,
    V_ql,
    V_qpe,
    V_iqpe,
    sheet_name="data"
):

    run_id = file_name   # run_id lấy trực tiếp từ file_name

    # ----- chuyển ma trận thành dataframe -----
    def pack(method, M):
        rows = []
        N, k = M.shape

        for j in range(k):
            row = {
                "run_id": run_id,
                "method": method,
                "eig_idx": j + 1
            }

            for i in range(N):
                row[f"v{i}"] = float(M[i, j])

            rows.append(row)

        return rows

    rows = []
    rows += pack("Traditional", vecs)
    rows += pack("QLanczos", V_ql)
    rows += pack("QPE", V_qpe)
    rows += pack("IQPE", V_iqpe)

    df = pd.DataFrame(rows)

    # ----- nếu file chưa tồn tại → tạo mới -----
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ----- nếu file tồn tại → append -----
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )

    print("✅ Appended data to:", excel_path)

def append_eigenvalues_simple(
    excel_path,
    file_name,
    evals,
    E_ql,
    E_qpe,
    E_iqpe,
    sheet_name="data"
):

    run_id = file_name

    def pack(method, E):
        row = {
            "run_id": run_id,
            "method": method
        }

        for i, val in enumerate(E):
            row[f"e{i}"] = float(val)

        return row

    rows = [
        pack("Traditional", evals),
        pack("QLanczos", E_ql),
        pack("QPE", E_qpe),
        pack("IQPE", E_iqpe)
    ]

    df = pd.DataFrame(rows)

    # ---------- nếu file chưa tồn tại ----------
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ---------- nếu file tồn tại ----------
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,   # luôn append
            startrow=startrow
        )

    print("✅ Appended eigenvalues to:", excel_path)

def append_log_excel(
        excel_path,
        file_name,
        start_vecs,
        end_vecs,
        start_V_ql,
        end_V_ql,
        end_V_qpe,
        end_V_iqpe,
        sheet_name="log"
    ):

    # ----- tạo dataframe -----
    new_df = pd.DataFrame(
        [(file_name, start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe)],
        columns=[
            "Tên file",
            "start_L",
            "end_L",
            "start_QL",
            "end_QL",
            "end_QPE",
            "end_IQPE"
        ]
    )

    # ----- nếu file chưa tồn tại → tạo mới -----
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            new_df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ----- file đã tồn tại -----
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        new_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )

    print(f"📝 Appended data to: {excel_path}")

def smallest_eigenpairs_ncut(W_coo, k):
    if not isinstance(W_coo, coo_matrix):
        W_coo = W_coo.tocoo()

    # chuẩn hóa
    D_vals = np.array(W_coo.sum(axis=1)).flatten()
    D_inv_sqrt = 1.0 / np.sqrt(D_vals + 1e-8)

    row, col = W_coo.row, W_coo.col
    data = W_coo.data * D_inv_sqrt[row] * D_inv_sqrt[col]
    W_norm = coo_matrix((data, (row, col)), shape=W_coo.shape)

    # A = I - W_norm
    n = W_coo.shape[0]
    A = coo_matrix(np.eye(n)) - W_norm

    # tìm k trị riêng nhỏ nhất
    evals, evecs = eigsh(A, k=k, which="SA", maxiter=5000)
    idx = np.argsort(evals)
    evals = evals[idx]
    evecs = evecs[:, idx]

    return evals, evecs

def main():

    excel_path = os.path.join("/content/drive/MyDrive/Test/log.xlsx")  # file Excel lưu
    excel_path_Vector = os.path.join("/content/drive/MyDrive/Test/logVector.xlsx")  # file Excel lưu
    excel_path_Eig = os.path.join("/content/drive/MyDrive/Test/logEig.xlsx")  # file Excel lưu
    input_path = "/content/drive/MyDrive/Test/in1"
    output_path = "/content/drive/MyDrive/Test/out"

    if not os.path.isdir(input_path):
        print(f"❌ Thư mục {input_path} không tồn tại!")
        exit()

    image_files = [f for f in os.listdir(input_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]
    if not image_files:
        print(f"❌ Không tìm thấy file ảnh nào trong {input_path}!")
        exit()

    log_rows = []  # mỗi phần tử: (tên file, bắt đầu, kết thúc)

    for idx, file_name in enumerate(image_files, start=1):
        start = time.perf_counter()
        k = int(re.search(r"_(\d+)\.png$", file_name).group(1))

        image_path = os.path.join(input_path, file_name)
        print(f"📷 Đang xử lý ảnh {idx}: {image_path}")

        sigma_i = 0.009
        sigma_x = 8

        save_image_name = os.path.join(output_path, f"{os.path.splitext(file_name)[0]}")
        start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe, evals, E_ql, E_qpe, E_iqpe, vecs, V_ql, V_qpe, V_iqpe = normalized_cuts_eigsh(file_name, image_path, save_image_name, k, sigma_i, sigma_x)
        
        append_log_excel(excel_path, file_name, start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe)
        append_eigenvalues_simple(excel_path_Eig, file_name, evals, E_ql, E_qpe, E_iqpe)
        append_eigenvectors_row_format(excel_path_Vector, file_name, vecs, V_ql, V_qpe, V_iqpe)
        end = time.perf_counter()
        print("Thời gian xử lý 1 ảnh ",end-start)

if __name__ == "__main__":
    main()
