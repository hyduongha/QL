import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
from datetime import datetime
import time
import numpy as np
import scipy.sparse as sp
from sklearn.cluster import KMeans
from skimage import io, color
import os
from sklearn.neighbors import NearestNeighbors
import pandas as pd
import re
from openpyxl import load_workbook
from scipy.linalg import expm, eigh
from scipy.sparse import coo_matrix
from qiskit import QuantumCircuit, transpile
from qiskit_aer import AerSimulator
from qiskit.circuit.library import UnitaryGate, QFT
from qiskit.quantum_info import Statevector
import argparse
# ============================================================
# Utils
# ============================================================
def _normalize(v, eps=1e-12):
    v = np.asarray(v, dtype=complex)
    n = np.linalg.norm(v)
    if n < eps:
        return v, 0.0
    return v / n, n

def _overlap_abs(v1, v2):
    v1 = np.asarray(v1, dtype=complex)
    v2 = np.asarray(v2, dtype=complex)
    n1 = np.linalg.norm(v1); n2 = np.linalg.norm(v2)
    if n1 < 1e-15 or n2 < 1e-15:
        return 0.0
    return float(np.abs(np.vdot(v1 / n1, v2 / n2)))

def _ql_vector_to_real(v_est: np.ndarray) -> np.ndarray:
    """
    Rotate complex vector by a global phase to maximize its real part,
    then return normalized real vector (for KMeans).
    """
    v_est = np.asarray(v_est, dtype=complex)
    v_est, nrm = _normalize(v_est)
    if nrm == 0:
        return np.real(v_est)

    a = np.real(v_est)
    b = np.imag(v_est)
    aa = float(a @ a); bb = float(b @ b); ab = float(a @ b)

    theta = 0.5 * np.arctan2(2.0 * ab, (aa - bb))
    v_rot = v_est * np.exp(-1j * theta)

    v_real = np.real(v_rot)
    nrm2 = np.linalg.norm(v_real)
    if nrm2 < 1e-15:
        return v_real
    return v_real / nrm2

def _deflate_against(v: np.ndarray, basis_vecs):
    """Project v to orthogonal complement of span(basis_vecs)."""
    vv = np.asarray(v, dtype=complex).copy()
    for b in (basis_vecs or []):
        b = np.asarray(b, dtype=complex)
        denom = np.vdot(b, b)
        if np.abs(denom) > 1e-15:
            vv = vv - (np.vdot(b, vv) / denom) * b
    vv, _ = _normalize(vv)
    return vv

# ============================================================
# QLanczos (Aer statevector Krylov, QR basis)
# ============================================================
def _apply_power_U_circuit(n_qubits: int, U: np.ndarray, power: int, init_state: np.ndarray) -> QuantumCircuit:
    qc = QuantumCircuit(n_qubits)
    qc.initialize(init_state.astype(complex), range(n_qubits))
    U_gate = UnitaryGate(U)
    for _ in range(power):
        qc.append(U_gate, range(n_qubits))
    return qc

def _sv_from_circuit(qc: QuantumCircuit, backend: AerSimulator) -> np.ndarray:
    tqc = transpile(qc, backend, optimization_level=0)
    tqc.save_statevector()
    result = backend.run(tqc, shots=1).result()
    sv = result.get_statevector(tqc)
    return np.asarray(sv, dtype=complex)

def _qlanczos_aer_qr_single_start(
    H: np.ndarray,
    k: int,
    m_krylov: int,
    t_evol: float,
    init_state: np.ndarray,
    reg_diag: float = 1e-12,
):
    """
    Single-start QLanczos with QR-orthonormalized Krylov basis.
    Returns (evals_k, evecs_k_fullspace) with evecs shape (N, k_eff).
    """
    H = np.asarray(H, dtype=complex)
    N = H.shape[0]

    if not np.allclose(H, H.conj().T, atol=1e-10):
        raise ValueError("H must be Hermitian/symmetric.")

    n_qubits = int(np.log2(N))
    if 2**n_qubits != N:
        raise ValueError(f"N={N} must be power of 2 (e.g., 32 for 4x8 image).")

    init_state, n0 = _normalize(init_state)
    if n0 == 0:
        raise ValueError("init_state is zero vector")

    U = expm(-1j * H * t_evol)
    backend = AerSimulator(method="statevector")

    states = []
    for j in range(m_krylov):
        qc = _apply_power_U_circuit(n_qubits, U, power=j, init_state=init_state)
        psi_j = _sv_from_circuit(qc, backend)
        psi_j, _ = _normalize(psi_j)
        states.append(psi_j)

    Psi = np.column_stack(states)  # (N, m_krylov)

    Q, R = np.linalg.qr(Psi)
    diagR = np.abs(np.diag(R))
    keep = diagR > reg_diag
    Qk = Q[:, keep] if np.any(keep) else Q

    Hproj = Qk.conj().T @ (H @ Qk)
    Hproj = 0.5 * (Hproj + Hproj.conj().T)

    evals_sub, evecs_sub = eigh(Hproj)  # ascending
    k_eff = min(k, len(evals_sub))
    evals_k = np.real(evals_sub[:k_eff])

    evecs_k = []
    for i in range(k_eff):
        y = evecs_sub[:, i]
        v = Qk @ y
        v, _ = _normalize(v)
        evecs_k.append(v)

    V = np.column_stack(evecs_k) if k_eff > 0 else np.zeros((N, 0), dtype=complex)
    return evals_k, V

def _ql_multistart_select(
    H: np.ndarray,
    k: int,
    n_starts: int,
    k_per_start: int,
    m_krylov: int,
    t_evol: float,
    seed: int,
    overlap_tol: float,
):
    """
    Multi-start: collect candidates, select k smallest distinct by overlap.
    Returns (E_sel, V_sel) where V_sel shape (N, k_found).
    """
    H = np.asarray(H, dtype=float)
    N = H.shape[0]
    rng = np.random.default_rng(seed)

    candidates = []
    for s in range(n_starts):
        init = rng.normal(size=N) + 1j * rng.normal(size=N)
        init, _ = _normalize(init)

        Es, Vs = _qlanczos_aer_qr_single_start(
            H, k=k_per_start, m_krylov=m_krylov, t_evol=t_evol, init_state=init
        )

        for j in range(Vs.shape[1]):
            vj, _ = _normalize(Vs[:, j])
            candidates.append({"E": float(Es[j]), "v": vj})

    candidates.sort(key=lambda x: x["E"])

    selected = []
    for cand in candidates:
        v = cand["v"]
        if all(_overlap_abs(v, sel["v"]) <= overlap_tol for sel in selected):
            selected.append(cand)
        if len(selected) >= k:
            break

    E_sel = np.array([s["E"] for s in selected], dtype=float)
    V_sel = np.column_stack([s["v"] for s in selected]) if selected else np.zeros((N, 0), dtype=complex)
    for i in range(V_sel.shape[1]):
        V_sel[:, i], _ = _normalize(V_sel[:, i])

    return E_sel, V_sel


# ============================================================
# QPE helpers (with notebook endianness fix)
# ============================================================
def _wrap_period(x, period):
    y = (x + 0.5 * period) % period - 0.5 * period
    return float(y)

def _energy_from_phase(phi, t_evol):
    ang = 2.0 * np.pi * float(phi)
    Ew = -(ang / t_evol)
    period = 2.0 * np.pi / t_evol
    return _wrap_period(Ew, period)

def _unwrap_to_reference(E_wrapped, E_ref, t_evol):
    period = 2.0 * np.pi / t_evol
    m = int(np.round((E_ref - E_wrapped) / period))
    return float(E_wrapped + m * period)

def _build_qpe_circuit_power(U: np.ndarray, t_count: int, psi_in: Statevector) -> QuantumCircuit:
    N = U.shape[0]
    n_state = int(np.log2(N))

    qc = QuantumCircuit(t_count + n_state)

    for q in range(t_count):
        qc.h(q)

    qc.initialize(psi_in.data, range(t_count, t_count + n_state))

    U_gate = UnitaryGate(U)

    for j in range(t_count):
        power = 2**j
        Ug = U_gate.power(power)
        cUg = Ug.control(1)
        qc.append(cUg, [j] + list(range(t_count, t_count + n_state)))

    qft_inv = QFT(t_count).inverse()
    qc.append(qft_inv, range(t_count))

    return qc

def _counting_probs(sv: Statevector, t_count: int):
    data = np.asarray(sv.data, dtype=complex)
    Ntot = data.shape[0]
    n_state = int(np.log2(Ntot)) - t_count
    amp = data.reshape((2**n_state, 2**t_count))
    probs = np.sum(np.abs(amp)**2, axis=0)
    out = {}
    for m in range(2**t_count):
        b = format(m, f"0{t_count}b")
        out[b] = float(probs[m])
    return out

def _bitstring_to_counting_index(bitstring: str, p_target: float, sv: Statevector, t_count: int, n_state: int) -> int:
    m1 = int(bitstring, 2)
    m2 = int(bitstring[::-1], 2)

    amp = np.asarray(sv.data, dtype=complex).reshape((2**n_state, 2**t_count))
    probs = np.sum(np.abs(amp)**2, axis=0)

    p1 = float(probs[m1])
    p2 = float(probs[m2])

    return m1 if abs(p1 - p_target) <= abs(p2 - p_target) else m2

def _qpe_postselect_best_mode(
    U: np.ndarray,
    t_count: int,
    psi_in: Statevector,
    E_target: float,
    t_evol: float,
    prev_refined_vecs=None,
    topM: int = 16,
    w_energy: float = 1.0,
    w_overlap: float = 0.25,
    w_dup: float = 5.0,
    min_mode_prob: float = 1e-6,
):
    prev_refined_vecs = prev_refined_vecs or []

    qc = _build_qpe_circuit_power(U, t_count, psi_in)
    sv = Statevector.from_instruction(qc)

    N = U.shape[0]
    n_state = int(np.log2(N))
    probs_count = _counting_probs(sv, t_count)
    top = sorted(probs_count.items(), key=lambda kv: kv[1], reverse=True)[:topM]

    amp = np.asarray(sv.data, dtype=complex).reshape((2**n_state, 2**t_count))

    best_score = None
    best_v = None
    best_info = None

    psi_vec = np.asarray(psi_in.data, dtype=complex)
    psi_vec, _ = _normalize(psi_vec)

    for bitstring, p_mode in top:
        p_mode = float(p_mode)
        if p_mode < min_mode_prob:
            continue

        m_mode = _bitstring_to_counting_index(bitstring, p_mode, sv, t_count, n_state)

        block = amp[:, m_mode]
        norm = np.linalg.norm(block)
        if norm < 1e-15:
            continue
        v = block / norm

        phi = int(bitstring, 2) / (2 ** t_count)
        E_wrapped = _energy_from_phase(phi, t_evol)
        E_refined = _unwrap_to_reference(E_wrapped, E_target, t_evol)

        ov_in = _overlap_abs(psi_vec, v)
        dup = 0.0
        if prev_refined_vecs:
            dup = max(_overlap_abs(v, vv) for vv in prev_refined_vecs)

        score = (w_energy * abs(E_refined - E_target)) - (w_overlap * ov_in) + (w_dup * dup)

        info = {
            "bit_mode": bitstring,
            "p_mode": p_mode,
            "m_mode_used": int(m_mode),
            "phi_mode": float(phi),
            "E_wrapped": float(E_wrapped),
            "E_refined": float(E_refined),
            "overlap_in": float(ov_in),
            "dup_overlap": float(dup),
            "score": float(score),
            "fallback": False,
        }

        if best_score is None or score < best_score:
            best_score = score
            best_info = info
            best_v = v

    if best_info is None:
        return None, {"fallback": True, "reason": "no_valid_mode_found"}

    best_v, _ = _normalize(best_v)
    return best_v, best_info


# ============================================================
# IQPE helper (Kitaev-style via inner products)
# ============================================================
def _precompute_U_powers(U: np.ndarray, n_powers: int):
    Upows = []
    cur = U.copy()
    Upows.append(cur)
    for _ in range(1, n_powers):
        cur = cur @ cur
        Upows.append(cur)
    return Upows

def _iqpe_kitaev_refine_energy(
    Upow_list,
    psi: np.ndarray,
    t_evol: float,
    E_target: float,
    weights: str = "exp",
):
    psi, n0 = _normalize(psi)
    if n0 == 0:
        return None, {"fallback": True, "reason": "zero_input"}

    E_list = []
    ang_list = []
    for k, U2k in enumerate(Upow_list):
        v = U2k @ psi
        z = np.vdot(psi, v)
        ang = np.angle(z)
        E_wrapped_k = -(ang) / (t_evol * (2**k))
        period_k = (2.0 * np.pi) / (t_evol * (2**k))
        E_wrapped_k = _wrap_period(E_wrapped_k, period_k)

        m = int(np.round((E_target - E_wrapped_k) / period_k))
        E_unwrapped_k = E_wrapped_k + m * period_k

        E_list.append(E_unwrapped_k)
        ang_list.append(ang)

    E_list = np.array(E_list, dtype=float)

    if weights == "uniform":
        w = np.ones_like(E_list)
    else:
        w = np.array([2**k for k in range(len(E_list))], dtype=float)

    E_refined = float(np.sum(w * E_list) / np.sum(w))
    return E_refined, {
        "fallback": False,
        "E_candidates": E_list,
        "angles": np.array(ang_list, dtype=float),
        "weights": w,
    }


# ============================================================
# MAIN PIPELINE for NCut: returns 3 sets of k eigenvectors
# ============================================================
def compute_ncut_ql_pipeline_three(
    A: np.ndarray,
    k: int,
    *,
    n_starts: int = 64,
    k_per_start: int = 8,
    m_krylov: int = 48,
    t_evol: float = 0.08,
    overlap_tol: float = 0.75,
    seed: int = 123,
    zero_tol: float = 1e-5,
    t_count: int = 10,
    topM: int = 16,
    w_energy: float = 1.0,
    w_overlap: float = 0.25,
    w_dup: float = 5.0,
    iqpe_powers: int = 10,
    iqpe_weights: str = "exp",
):
    start_V_ql = time.perf_counter()
    
    A = np.asarray(A, dtype=float)
    N = A.shape[0]
    
    n_qubits = int(np.log2(N))
    if 2**n_qubits != N:
        raise ValueError(f"N={N} must be power of 2 for Aer statevector QLanczos/QPE/IQPE.")

    E1, V1 = _ql_multistart_select(
        A,
        k=max(k + 3, k),
        n_starts=n_starts,
        k_per_start=k_per_start,
        m_krylov=m_krylov,
        t_evol=t_evol,
        seed=seed,
        overlap_tol=overlap_tol,
    )
    if V1.shape[1] == 0:
        raise RuntimeError("Step1 QLanczos produced no candidates.")

    idx = np.argsort(E1)
    E1 = E1[idx]
    V1 = V1[:, idx]

    ###################### Thay bằng 3 dòng này ##########################
    k_eff = min(k, len(E1))
    E_ql = E1[1:k_eff+1]
    V_ql_c = V1[:, 1:k_eff+1]
    '''
    keep = np.where(np.abs(E1) > zero_tol)[0]
    if keep.size == 0:
        keep = np.arange(min(k, len(E1)))
    keep = keep[:k]

    E_ql = E1[keep]
    V_ql_c = V1[:, keep]
    '''
    ###################### Thay bằng 3 dòng này ##########################

    
    V_ql = np.column_stack([_ql_vector_to_real(V_ql_c[:, i]) for i in range(V_ql_c.shape[1])]).astype(np.float32)

    U = expm(-1j * A * t_evol)

    end_V_ql = time.perf_counter()

    V_qpe_list = []
    E_qpe_list = []
    refined_vecs_for_dup = []

    k_eff = V_ql_c.shape[1]
    for i in range(k_eff):
        psi = V_ql_c[:, i]
        psi = _deflate_against(psi, refined_vecs_for_dup)
        psi_sv = Statevector(psi)

        v_ref, info = _qpe_postselect_best_mode(
            U=U,
            t_count=t_count,
            psi_in=psi_sv,
            E_target=float(E_ql[i]),
            t_evol=t_evol,
            prev_refined_vecs=refined_vecs_for_dup,
            topM=topM,
            w_energy=w_energy,
            w_overlap=w_overlap,
            w_dup=w_dup,
        )

        if v_ref is None:
            v_ref = psi
            E_qpe_list.append(float(E_ql[i]))
        else:
            E_qpe_list.append(float(info["E_refined"]))

        v_ref, _ = _normalize(v_ref)
        refined_vecs_for_dup.append(v_ref)
        V_qpe_list.append(_ql_vector_to_real(v_ref))

    V_qpe = np.column_stack(V_qpe_list).astype(np.float32)
    E_qpe = np.array(E_qpe_list, dtype=float)

    end_V_qpe = time.perf_counter()

    Upow_list = _precompute_U_powers(U, iqpe_powers)

    V_iqpe_list = []
    E_iqpe_list = []
    iqpe_selected_vecs = []

    k_eff = V_ql_c.shape[1]
    for i in range(k_eff):
        psi = V_ql_c[:, i]
        psi = _deflate_against(psi, iqpe_selected_vecs)
        psi, _ = _normalize(psi)

        E_refined, _info = _iqpe_kitaev_refine_energy(
            Upow_list=Upow_list,
            psi=psi,
            t_evol=t_evol,
            E_target=float(E_ql[i]),
            weights=iqpe_weights,
        )

        if E_refined is None:
            E_iqpe_list.append(float(E_ql[i]))
        else:
            E_iqpe_list.append(float(E_refined))

        iqpe_selected_vecs.append(psi)
        V_iqpe_list.append(_ql_vector_to_real(psi))

    V_iqpe = np.column_stack(V_iqpe_list).astype(np.float32)
    E_iqpe = np.array(E_iqpe_list, dtype=float)

    end_V_iqpe = time.perf_counter()

    return E_ql, E_qpe, E_iqpe, V_ql, V_qpe, V_iqpe, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe

def compute_weight_matrix_coo_knn(image, sigma_i, sigma_x, k_neighbors=10):
    h, w, c = image.shape
    N = h * w

    coords = np.array(np.meshgrid(range(h), range(w))).reshape(2, -1).T
    features = image.reshape(-1, c)

    knn = NearestNeighbors(n_neighbors=k_neighbors, algorithm="ball_tree")
    knn.fit(coords)
    distances, indices = knn.kneighbors(coords)

    row_idx = np.repeat(np.arange(N), k_neighbors)
    col_idx = indices.flatten()

    feat_row = features[row_idx]
    feat_col = features[col_idx]

    diff = feat_row - feat_col
    W_feature = np.exp(-np.sum(diff ** 2, axis=1) / (2 * sigma_i ** 2))
    W_space = np.exp(-(distances.flatten() ** 2) / (2 * sigma_x ** 2))

    values = W_feature * W_space

    W_sparse = sp.coo_matrix((values, (row_idx, col_idx)), shape=(N, N))
    return W_sparse


def compute_laplacian_coo(W_coo):
    D = np.array(W_coo.sum(axis=1)).flatten()
    D_coo = sp.diags(D)
    L_coo = D_coo - W_coo
    return L_coo, D_coo

def assign_labels(eigen_vectors, k):
    return KMeans(n_clusters=k, random_state=0).fit(eigen_vectors).labels_

def save_segmentation(image, labels, k, output_path):
    h, w, c = image.shape
    segmented_image = np.zeros_like(image, dtype=np.uint8)
    for i in range(k):
        mask = labels.reshape(h, w) == i
        cluster_pixels = image[mask]
        mean_color = (cluster_pixels.mean(axis=0) * 255).astype(np.uint8) if len(cluster_pixels) > 0 else np.array([0, 0, 0], dtype=np.uint8)
        segmented_image[mask] = mean_color
    io.imsave(output_path, segmented_image)

def save_seg_file(labels, image_shape, output_path, image_name="image"):
    h, w = image_shape[:2]
    unique_labels = np.unique(labels)
    segments = len(unique_labels)

    # Tạo phần header
    header = [
        "format ascii cr",
        f"date {datetime.now().strftime('%a %b %d %H:%M:%S %Y')}",
        f"image ",
        "user 1102",  # Giữ nguyên như file mẫu
        f"width {w}",
        f"height {h}",
        f"segments {segments}",
        "gray 0",
        "invert 0",
        "flipflop 0",
        "data"
    ]

    # Tạo dữ liệu pixel theo định dạng (nhãn, dòng, cột bắt đầu, cột kết thúc)
    data_lines = []
    for row in range(h):
        row_labels = labels[row, :]
        start_col = 0
        current_label = row_labels[0]

        for col in range(1, w):
            if row_labels[col] != current_label:
                data_lines.append(f"{current_label} {row} {start_col} {col}")
                start_col = col
                current_label = row_labels[col]

        # Thêm dòng cuối cùng của hàng
        data_lines.append(f"{current_label} {row} {start_col} {w}")

    # Lưu vào file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(header) + "\n")
        f.write("\n".join(data_lines) + "\n")

    print(f"✅ File SEG đã lưu: {output_path}")

def align_eigenvector_signs(vecs, V_ql, V_qpe, V_iqpe):
    """
    Align signs of V_ql, V_qpe, V_iqpe to match reference eigenvectors vecs.

    Parameters
    ----------
    vecs : np.ndarray (N,k)
        Reference eigenvectors (traditional Lanczos).
    V_ql : np.ndarray (N,k)
    V_qpe : np.ndarray (N,k)
    V_iqpe : np.ndarray (N,k)

    Returns
    -------
    V_ql_aligned, V_qpe_aligned, V_iqpe_aligned
    """

    k = vecs.shape[1]

    V_ql_aligned = V_ql.copy()
    V_qpe_aligned = V_qpe.copy()
    V_iqpe_aligned = V_iqpe.copy()

    for i in range(k):
        ref = vecs[:, i]

        if np.dot(ref, V_ql_aligned[:, i]) < 0:
            V_ql_aligned[:, i] *= -1

        if np.dot(ref, V_qpe_aligned[:, i]) < 0:
            V_qpe_aligned[:, i] *= -1

        if np.dot(ref, V_iqpe_aligned[:, i]) < 0:
            V_iqpe_aligned[:, i] *= -1

    return V_ql_aligned, V_qpe_aligned, V_iqpe_aligned

def build_ncut_matrix(W_coo):
    if not isinstance(W_coo, coo_matrix):
        W_coo = W_coo.tocoo()

    W_sym = (W_coo + W_coo.T).tocoo()
    W_sym.data *= 0.5

    D_vals = np.array(W_sym.sum(axis=1)).flatten()
    D_inv_sqrt = 1.0 / np.sqrt(D_vals + 1e-8)

    row, col = W_sym.row, W_sym.col
    data = W_sym.data * D_inv_sqrt[row] * D_inv_sqrt[col]
    W_norm = coo_matrix((data, (row, col)), shape=W_sym.shape).toarray()

    A = np.eye(W_coo.shape[0], dtype=float) - W_norm
    A = 0.5 * (A + A.T)
    return A
    
def row_normalize(X, eps=1e-12):
    X = np.asarray(X, dtype=float)
    norms = np.linalg.norm(X, axis=1, keepdims=True)
    return X / np.maximum(norms, eps)
    
def normalized_cuts_eigsh(imagename, image_path, output_path, k, sigma_i, sigma_x):
    image = io.imread(image_path)
    image = color.gray2rgb(image) if image.ndim == 2 else image[:, :, :3] if image.shape[2] == 4 else image
    image = image / 255.0

    start_vecs = time.perf_counter()
    W_coo = compute_weight_matrix_coo_knn(image, sigma_i, sigma_x)
    A = build_ncut_matrix(W_coo)
    A = np.asarray(A, dtype=float)
    evals, vecs = np.linalg.eigh(A)
    idx = np.argsort(evals)[1:k+1]
    evals = evals[idx]
    vecs = vecs[:, idx]
    
    end_vecs = time.perf_counter()

    E_ql, E_qpe, E_iqpe, V_ql, V_qpe, V_iqpe, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe = compute_ncut_ql_pipeline_three(A, k)

    V_ql, V_qpe, V_iqpe = align_eigenvector_signs(vecs, V_ql, V_qpe, V_iqpe)

    labels = assign_labels(row_normalize(vecs), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_L.seg", imagename)

    labels = assign_labels(row_normalize(V_ql), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_QL.seg", imagename)

    labels = assign_labels(row_normalize(V_qpe), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_QPE.seg", imagename)

    labels = assign_labels(row_normalize(V_iqpe), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_IQPE.seg", imagename)
    
    del W_coo

    return start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe, evals, E_ql, E_qpe, E_iqpe, vecs, V_ql, V_qpe, V_iqpe

def append_eigenvectors_row_format(
    excel_path,
    file_name,
    vecs,
    V_ql,
    V_qpe,
    V_iqpe,
    sheet_name="data"
):

    run_id = file_name   # run_id lấy trực tiếp từ file_name

    # ----- chuyển ma trận thành dataframe -----
    def pack(method, M):
        rows = []
        N, k = M.shape

        for j in range(k):
            row = {
                "run_id": run_id,
                "method": method,
                "eig_idx": j + 1
            }

            for i in range(N):
                row[f"v{i}"] = float(np.real(M[i, j]))

            rows.append(row)

        return rows

    rows = []
    rows += pack("Traditional", vecs)
    rows += pack("QLanczos", V_ql)
    rows += pack("QPE", V_qpe)
    rows += pack("IQPE", V_iqpe)

    df = pd.DataFrame(rows)

    # ----- nếu file chưa tồn tại → tạo mới -----
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ----- nếu file tồn tại → append -----
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )

    print("✅ Appended data to:", excel_path)

def append_eigenvalues_simple(
    excel_path,
    file_name,
    evals,
    E_ql,
    E_qpe,
    E_iqpe,
    sheet_name="data"
):

    run_id = file_name

    def pack(method, E):
        row = {
            "run_id": run_id,
            "method": method
        }

        for i, val in enumerate(E):
            row[f"e{i}"] = float(val)

        return row

    rows = [
        pack("Traditional", evals),
        pack("QLanczos", E_ql),
        pack("QPE", E_qpe),
        pack("IQPE", E_iqpe)
    ]

    df = pd.DataFrame(rows)

    # ---------- nếu file chưa tồn tại ----------
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ---------- nếu file tồn tại ----------
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,   # luôn append
            startrow=startrow
        )

    print("✅ Appended eigenvalues to:", excel_path)

def append_log_excel(
        excel_path,
        file_name,
        start_vecs,
        end_vecs,
        start_V_ql,
        end_V_ql,
        end_V_qpe,
        end_V_iqpe,
        sheet_name="log"
    ):

    # ----- tạo dataframe -----
    new_df = pd.DataFrame(
        [(file_name, start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe)],
        columns=[
            "Tên file",
            "start_L",
            "end_L",
            "start_QL",
            "end_QL",
            "end_QPE",
            "end_IQPE"
        ]
    )

    # ----- nếu file chưa tồn tại → tạo mới -----
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            new_df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ----- file đã tồn tại -----
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        new_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )

    print(f"📝 Appended data to: {excel_path}")

def main(name):
    input_path = "/content/drive/MyDrive/Test/" + name
    excel_path = os.path.join("/content/drive/MyDrive/Test/log.xlsx")  # file Excel lưu
    excel_path_Vector = os.path.join("/content/drive/MyDrive/Test/logVector.xlsx")  # file Excel lưu
    excel_path_Eig = os.path.join("/content/drive/MyDrive/Test/logEig.xlsx")  # file Excel lưu
    output_path = "/content/drive/MyDrive/Test/out"

    if not os.path.isdir(input_path):
        print(f"❌ Thư mục {input_path} không tồn tại!")
        exit()
    os.makedirs(output_path, exist_ok=True)
    
    image_files = [f for f in os.listdir(input_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]
    if not image_files:
        print(f"❌ Không tìm thấy file ảnh nào trong {input_path}!")
        exit()

    log_rows = []  # mỗi phần tử: (tên file, bắt đầu, kết thúc)

    for idx, file_name in enumerate(image_files, start=1):
        start = time.perf_counter()
        k = int(re.search(r"_(\d+)\.png$", file_name).group(1))

        image_path = os.path.join(input_path, file_name)
        print(f"📷 Đang xử lý ảnh {idx}: {image_path}")

        sigma_i = 0.009
        sigma_x = 8

        save_image_name = os.path.join(output_path, f"{os.path.splitext(file_name)[0]}")
        start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe, evals, E_ql, E_qpe, E_iqpe, vecs, V_ql, V_qpe, V_iqpe = normalized_cuts_eigsh(file_name, image_path, save_image_name, k, sigma_i, sigma_x)
        
        append_log_excel(excel_path, file_name, start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe)
        append_eigenvalues_simple(excel_path_Eig, file_name, evals, E_ql, E_qpe, E_iqpe)
        append_eigenvectors_row_format(excel_path_Vector, file_name, vecs, V_ql, V_qpe, V_iqpe)
        end = time.perf_counter()
        print("Thời gian xử lý 1 ảnh ",end-start)

#####################################      Chỗ sửa đường dẫn     ######################################################
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("nameImage", type=str)
    args = parser.parse_args()
    main(args.nameImage)
########################################################################################################################
