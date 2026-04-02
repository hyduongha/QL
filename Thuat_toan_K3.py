import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
from datetime import datetime
import time
import numpy as np
import scipy.sparse as sp
from sklearn.cluster import KMeans
from skimage import io, color
import os
from sklearn.neighbors import NearestNeighbors
import pandas as pd
import re
from openpyxl import load_workbook
from scipy.linalg import expm, eigh
from scipy.sparse import coo_matrix
from qiskit import QuantumCircuit, transpile
from qiskit_aer import AerSimulator
from qiskit.circuit.library import UnitaryGate, QFT
from qiskit.quantum_info import Statevector
import argparse
# ============================================================
# Utils
# ============================================================
def _residual_norm(A: np.ndarray, lam: float, v: np.ndarray) -> float:
    v = np.asarray(v, dtype=complex)
    return float(np.linalg.norm(A @ v - lam * v))


def _canonicalize_phase(v: np.ndarray) -> np.ndarray:
    """
    Fix global phase using the largest-magnitude entry,
    so vectors are more stable before taking real part.
    """
    v = np.asarray(v, dtype=complex)
    v, nrm = _normalize(v)
    if nrm == 0:
        return v
    idx = int(np.argmax(np.abs(v)))
    phase = np.angle(v[idx])
    v = v * np.exp(-1j * phase)
    return v


def _ritz_refine_from_candidates(
    A: np.ndarray,
    V_cand: np.ndarray,
    n_keep: int,
    qr_tol: float = 1e-12,
):
    """
    QL candidates -> QR -> Ritz refinement.
    Returns refined eigenpairs in ascending order.
    """
    if V_cand.size == 0:
        return np.array([], dtype=float), np.zeros((A.shape[0], 0), dtype=complex)

    V_cand = np.asarray(V_cand, dtype=complex)

    Q, R = np.linalg.qr(V_cand)
    diagR = np.abs(np.diag(R))
    keep = diagR > qr_tol
    Qk = Q[:, keep] if np.any(keep) else Q

    Hproj = Qk.conj().T @ (A @ Qk)
    Hproj = 0.5 * (Hproj + Hproj.conj().T)

    ew, Z = eigh(Hproj)
    idx = np.argsort(np.real(ew))
    ew = np.real(ew[idx])
    Z = Z[:, idx]

    n_keep = min(n_keep, len(ew))
    ew = ew[:n_keep]

    V_ref = []
    for i in range(n_keep):
        v = Qk @ Z[:, i]
        v = _canonicalize_phase(v)
        v, _ = _normalize(v)
        V_ref.append(v)

    V_ref = np.column_stack(V_ref) if len(V_ref) > 0 else np.zeros((A.shape[0], 0), dtype=complex)
    return ew, V_ref


def _select_nontrivial_by_degree(
    E: np.ndarray,
    V: np.ndarray,
    D_vals: np.ndarray,
    k: int,
):
    """
    Chọn k eigenvectors non-trivial cho normalized Laplacian.
    Vector trivial ~ sqrt(D).
    """
    E = np.asarray(E, dtype=float)
    V = np.asarray(V, dtype=complex)

    if V.shape[1] == 0:
        return np.array([], dtype=float), np.zeros((V.shape[0], 0), dtype=complex)

    trivial = np.sqrt(np.asarray(D_vals, dtype=float))
    trivial = trivial / max(np.linalg.norm(trivial), 1e-12)

    overlaps = []
    trivial_c = trivial.astype(complex)
    overlaps = []
    for i in range(V.shape[1]):
        vi = _canonicalize_phase(V[:, i])
        vi, _ = _normalize(vi)
        overlaps.append(float(abs(np.vdot(vi, trivial_c))))

    trivial_idx = int(np.argmax(overlaps))

    keep = [i for i in range(V.shape[1]) if i != trivial_idx]
    keep = sorted(keep, key=lambda i: E[i])[:k]

    E_sel = E[keep]
    V_sel = V[:, keep]

    return E_sel, V_sel


def _realify_columns(V: np.ndarray) -> np.ndarray:
    """
    Convert complex eigenvectors to real vectors only at the end.
    """
    if V.shape[1] == 0:
        return np.zeros((V.shape[0], 0), dtype=np.float32)

    cols = []
    for i in range(V.shape[1]):
        v = _canonicalize_phase(V[:, i])
        v = _ql_vector_to_real(v)
        cols.append(v.astype(np.float32))
    return np.column_stack(cols)

def _normalize(v, eps=1e-12):
    v = np.asarray(v, dtype=complex)
    n = np.linalg.norm(v)
    if n < eps:
        return v, 0.0
    return v / n, n

def _overlap_abs(v1, v2):
    v1 = np.asarray(v1, dtype=complex)
    v2 = np.asarray(v2, dtype=complex)
    n1 = np.linalg.norm(v1); n2 = np.linalg.norm(v2)
    if n1 < 1e-15 or n2 < 1e-15:
        return 0.0
    return float(np.abs(np.vdot(v1 / n1, v2 / n2)))

def _ql_vector_to_real(v_est: np.ndarray) -> np.ndarray:
    """
    Rotate complex vector by a global phase to maximize its real part,
    then return normalized real vector (for KMeans).
    """
    v_est = np.asarray(v_est, dtype=complex)
    v_est, nrm = _normalize(v_est)
    if nrm == 0:
        return np.real(v_est)

    a = np.real(v_est)
    b = np.imag(v_est)
    aa = float(a @ a); bb = float(b @ b); ab = float(a @ b)

    theta = 0.5 * np.arctan2(2.0 * ab, (aa - bb))
    v_rot = v_est * np.exp(-1j * theta)

    v_real = np.real(v_rot)
    nrm2 = np.linalg.norm(v_real)
    if nrm2 < 1e-15:
        return v_real
    return v_real / nrm2

def _deflate_against(v: np.ndarray, basis_vecs):
    """Project v to orthogonal complement of span(basis_vecs)."""
    vv = np.asarray(v, dtype=complex).copy()
    for b in (basis_vecs or []):
        b = np.asarray(b, dtype=complex)
        denom = np.vdot(b, b)
        if np.abs(denom) > 1e-15:
            vv = vv - (np.vdot(b, vv) / denom) * b
    vv, _ = _normalize(vv)
    return vv

# ============================================================
# QLanczos (Aer statevector Krylov, QR basis)
# ============================================================
def _apply_power_U_circuit(n_qubits: int, U: np.ndarray, power: int, init_state: np.ndarray) -> QuantumCircuit:
    qc = QuantumCircuit(n_qubits)
    qc.initialize(init_state.astype(complex), range(n_qubits))
    U_gate = UnitaryGate(U)
    for _ in range(power):
        qc.append(U_gate, range(n_qubits))
    return qc

def _sv_from_circuit(qc: QuantumCircuit, backend: AerSimulator) -> np.ndarray:
    tqc = transpile(qc, backend, optimization_level=0)
    tqc.save_statevector()
    result = backend.run(tqc, shots=1).result()
    sv = result.get_statevector(tqc)
    return np.asarray(sv, dtype=complex)

def _qlanczos_aer_qr_single_start(
    H: np.ndarray,
    k: int,
    m_krylov: int,
    t_evol: float,
    init_state: np.ndarray,
    reg_diag: float = 1e-12,
):
    """
    Single-start QLanczos with QR-orthonormalized Krylov basis.
    Returns (evals_k, evecs_k_fullspace) with evecs shape (N, k_eff).
    """
    H = np.asarray(H, dtype=complex)
    N = H.shape[0]

    if not np.allclose(H, H.conj().T, atol=1e-10):
        raise ValueError("H must be Hermitian/symmetric.")

    n_qubits = int(np.log2(N))
    if 2**n_qubits != N:
        raise ValueError(f"N={N} must be power of 2 (e.g., 32 for 4x8 image).")

    init_state, n0 = _normalize(init_state)
    if n0 == 0:
        raise ValueError("init_state is zero vector")

    U = expm(-1j * H * t_evol)
    backend = AerSimulator(method="statevector")

    states = []
    for j in range(m_krylov):
        qc = _apply_power_U_circuit(n_qubits, U, power=j, init_state=init_state)
        psi_j = _sv_from_circuit(qc, backend)
        psi_j, _ = _normalize(psi_j)
        states.append(psi_j)

    Psi = np.column_stack(states)  # (N, m_krylov)

    Q, R = np.linalg.qr(Psi)
    diagR = np.abs(np.diag(R))
    keep = diagR > reg_diag
    Qk = Q[:, keep] if np.any(keep) else Q

    Hproj = Qk.conj().T @ (H @ Qk)
    Hproj = 0.5 * (Hproj + Hproj.conj().T)

    evals_sub, evecs_sub = eigh(Hproj)  # ascending
    k_eff = min(k, len(evals_sub))
    evals_k = np.real(evals_sub[:k_eff])

    evecs_k = []
    for i in range(k_eff):
        y = evecs_sub[:, i]
        v = Qk @ y
        v, _ = _normalize(v)
        evecs_k.append(v)

    V = np.column_stack(evecs_k) if k_eff > 0 else np.zeros((N, 0), dtype=complex)
    return evals_k, V

def _ql_multistart_select(
    H: np.ndarray,
    k: int,
    n_starts: int,
    k_per_start: int,
    m_krylov: int,
    t_evol: float,
    seed: int,
    overlap_tol: float,
):
    """
    Multi-start: collect candidates, select distinct good candidates.
    Ưu tiên residual nhỏ rồi mới đến eigenvalue nhỏ.
    Returns (E_sel, V_sel) where V_sel shape (N, k_found).
    """
    H = np.asarray(H, dtype=float)
    N = H.shape[0]
    rng = np.random.default_rng(seed)

    candidates = []
    for s in range(n_starts):
        init = rng.normal(size=N) + 1j * rng.normal(size=N)
        init, _ = _normalize(init)

        Es, Vs = _qlanczos_aer_qr_single_start(
            H, k=k_per_start, m_krylov=m_krylov, t_evol=t_evol, init_state=init
        )

        for j in range(Vs.shape[1]):
            vj = _canonicalize_phase(Vs[:, j])
            vj, _ = _normalize(vj)

            lam = float(np.real(np.vdot(vj, H @ vj)))
            res = _residual_norm(H, lam, vj)

            candidates.append({
                "E": lam,
                "v": vj,
                "res": res
            })

    candidates.sort(key=lambda x: (x["E"], x["res"]))

    selected = []
    for cand in candidates:
        v = cand["v"]
        if all(_overlap_abs(v, sel["v"]) <= overlap_tol for sel in selected):
            selected.append(cand)
        if len(selected) >= k:
            break

    E_sel = np.array([s["E"] for s in selected], dtype=float)
    V_sel = np.column_stack([s["v"] for s in selected]) if selected else np.zeros((N, 0), dtype=complex)

    for i in range(V_sel.shape[1]):
        V_sel[:, i], _ = _normalize(V_sel[:, i])

    return E_sel, V_sel

# ============================================================
# QPE helpers (with notebook endianness fix)
# ============================================================
def _wrap_period(x, period):
    y = (x + 0.5 * period) % period - 0.5 * period
    return float(y)

def _energy_from_phase(phi, t_evol):
    ang = 2.0 * np.pi * float(phi)
    Ew = -(ang / t_evol)
    period = 2.0 * np.pi / t_evol
    return _wrap_period(Ew, period)

def _unwrap_to_reference(E_wrapped, E_ref, t_evol):
    period = 2.0 * np.pi / t_evol
    m = int(np.round((E_ref - E_wrapped) / period))
    return float(E_wrapped + m * period)

def _build_qpe_circuit_power(U: np.ndarray, t_count: int, psi_in: Statevector) -> QuantumCircuit:
    N = U.shape[0]
    n_state = int(np.log2(N))

    qc = QuantumCircuit(t_count + n_state)

    for q in range(t_count):
        qc.h(q)

    qc.initialize(psi_in.data, range(t_count, t_count + n_state))
    U = np.asarray(U, dtype=np.complex128)

    for j in range(t_count):
        power = 2**j

        # tính lũy thừa bằng numpy thay vì U_gate.power(...)
        U_pow = np.linalg.matrix_power(U, power)
        U_pow = np.asarray(U_pow, dtype=np.complex128)

        # dọn nhiễu số nhỏ
        U_pow.real[np.abs(U_pow.real) < 1e-14] = 0.0
        U_pow.imag[np.abs(U_pow.imag) < 1e-14] = 0.0

        Ug = UnitaryGate(U_pow)
        cUg = Ug.control(1)
        qc.append(cUg, [j] + list(range(t_count, t_count + n_state)))

    qft_inv = QFT(t_count).inverse()
    qc.append(qft_inv, range(t_count))

    return qc

def _counting_probs(sv: Statevector, t_count: int):
    data = np.asarray(sv.data, dtype=complex)
    Ntot = data.shape[0]
    n_state = int(np.log2(Ntot)) - t_count
    amp = data.reshape((2**n_state, 2**t_count))
    probs = np.sum(np.abs(amp)**2, axis=0)
    out = {}
    for m in range(2**t_count):
        b = format(m, f"0{t_count}b")
        out[b] = float(probs[m])
    return out

def _bitstring_to_counting_index(bitstring: str, p_target: float, sv: Statevector, t_count: int, n_state: int) -> int:
    m1 = int(bitstring, 2)
    m2 = int(bitstring[::-1], 2)

    amp = np.asarray(sv.data, dtype=complex).reshape((2**n_state, 2**t_count))
    probs = np.sum(np.abs(amp)**2, axis=0)

    p1 = float(probs[m1])
    p2 = float(probs[m2])

    return m1 if abs(p1 - p_target) <= abs(p2 - p_target) else m2

def _qpe_postselect_best_mode(
    U: np.ndarray,
    t_count: int,
    psi_in: Statevector,
    E_target: float,
    t_evol: float,
    prev_refined_vecs=None,
    topM: int = 16,
    w_energy: float = 1.0,
    w_overlap: float = 0.25,
    w_dup: float = 5.0,
    min_mode_prob: float = 1e-6,
):
    prev_refined_vecs = prev_refined_vecs or []

    qc = _build_qpe_circuit_power(U, t_count, psi_in)
    sv = Statevector.from_instruction(qc)

    N = U.shape[0]
    n_state = int(np.log2(N))
    probs_count = _counting_probs(sv, t_count)
    top = sorted(probs_count.items(), key=lambda kv: kv[1], reverse=True)[:topM]

    amp = np.asarray(sv.data, dtype=complex).reshape((2**n_state, 2**t_count))

    best_score = None
    best_v = None
    best_info = None

    psi_vec = np.asarray(psi_in.data, dtype=complex)
    psi_vec, _ = _normalize(psi_vec)

    for bitstring, p_mode in top:
        p_mode = float(p_mode)
        if p_mode < min_mode_prob:
            continue

        m_mode = _bitstring_to_counting_index(bitstring, p_mode, sv, t_count, n_state)

        block = amp[:, m_mode]
        norm = np.linalg.norm(block)
        if norm < 1e-15:
            continue
        v = block / norm

        phi = int(bitstring, 2) / (2 ** t_count)
        E_wrapped = _energy_from_phase(phi, t_evol)
        E_refined = _unwrap_to_reference(E_wrapped, E_target, t_evol)

        ov_in = _overlap_abs(psi_vec, v)
        dup = 0.0
        if prev_refined_vecs:
            dup = max(_overlap_abs(v, vv) for vv in prev_refined_vecs)

        score = (w_energy * abs(E_refined - E_target)) - (w_overlap * ov_in) + (w_dup * dup)

        info = {
            "bit_mode": bitstring,
            "p_mode": p_mode,
            "m_mode_used": int(m_mode),
            "phi_mode": float(phi),
            "E_wrapped": float(E_wrapped),
            "E_refined": float(E_refined),
            "overlap_in": float(ov_in),
            "dup_overlap": float(dup),
            "score": float(score),
            "fallback": False,
        }

        if best_score is None or score < best_score:
            best_score = score
            best_info = info
            best_v = v

    if best_info is None:
        return None, {"fallback": True, "reason": "no_valid_mode_found"}

    best_v, _ = _normalize(best_v)
    return best_v, best_info


# ============================================================
# IQPE helper (iterative / semiclassical, ancilla + controlled-U^(2^k))
# ============================================================
def _precompute_U_powers(U: np.ndarray, n_bits: int):
    """
    Upows[k] = U^(2^k)
    """
    Upows = [np.asarray(U, dtype=complex)]
    for _ in range(1, n_bits):
        Upows.append(Upows[-1] @ Upows[-1])
    return Upows


def _binary_fraction_from_tail_bits(bits_tail):
    """
    bits_tail = [x_{j+1}, x_{j+2}, ..., x_m]
    Trả về giá trị 0.0 x_{j+1} x_{j+2} ... x_m
    """
    frac = 0.0
    for idx, b in enumerate(bits_tail):
        frac += int(b) / (2 ** (idx + 2))
    return frac


def _build_iqpe_round_circuit(U_power: np.ndarray, psi: np.ndarray, feedback_frac: float) -> QuantumCircuit:
    """
    Một vòng IQPE:
      - 1 ancilla
      - initialize |psi> trên state register
      - H ancilla
      - controlled-U^(2^k)
      - phase feedback trên ancilla
      - H ancilla
      - measure ancilla
    """
    psi = np.asarray(psi, dtype=complex)
    psi, nrm = _normalize(psi)
    if nrm == 0:
        raise ValueError("psi is zero vector")

    U_power = np.asarray(U_power, dtype=np.complex128)
    # dọn nhiễu số nhỏ
    U_power.real[np.abs(U_power.real) < 1e-14] = 0.0
    U_power.imag[np.abs(U_power.imag) < 1e-14] = 0.0

    N = U_power.shape[0]
    n_state = int(np.log2(N))
    if 2 ** n_state != N:
        raise ValueError("State dimension must be power of 2")

    qc = QuantumCircuit(1 + n_state, 1)

    anc = 0
    state_qubits = list(range(1, 1 + n_state))

    qc.initialize(psi, state_qubits)
    qc.h(anc)

    U_gate = UnitaryGate(U_power)
    cU = U_gate.control(1)
    qc.append(cU, [anc] + state_qubits)

    # feedback angle = -2π * (0.0 tail_bits)
    if abs(feedback_frac) > 0:
        qc.p(-2.0 * np.pi * feedback_frac, anc)

    qc.h(anc)
    qc.measure(anc, 0)

    return qc


def _run_iqpe_round(
    backend: AerSimulator,
    U_power: np.ndarray,
    psi: np.ndarray,
    feedback_frac: float,
    shots: int = 1024,
    optimization_level: int = 0,
):
    qc = _build_iqpe_round_circuit(U_power, psi, feedback_frac)
    tqc = transpile(qc, backend, optimization_level=optimization_level)
    result = backend.run(tqc, shots=shots).result()
    counts = result.get_counts(tqc)

    c0 = int(counts.get("0", 0))
    c1 = int(counts.get("1", 0))

    # majority vote
    bit = 1 if c1 > c0 else 0

    return bit, {
        "counts": {"0": c0, "1": c1},
        "p0": c0 / max(shots, 1),
        "p1": c1 / max(shots, 1),
    }


def _bits_to_phase(bits_msb_first):
    """
    bits_msb_first = [x1, x2, ..., xm] với phi = 0.x1 x2 ... xm
    """
    phi = 0.0
    for i, b in enumerate(bits_msb_first):
        phi += int(b) / (2 ** (i + 1))
    return float(phi)


def _iqpe_iterative_refine_energy(
    U: np.ndarray,
    psi: np.ndarray,
    t_evol: float,
    E_target: float,
    n_bits: int = 10,
    shots: int = 1024,
    optimization_level: int = 0,
):
    """
    IQPE "thật" hơn:
      - 1 ancilla
      - controlled-U^(2^k)
      - feedback từ các bit đã đo
      - đo từng vòng để suy ra từng bit pha
    Trả về E_refined và log chi tiết từng vòng.
    """
    psi = np.asarray(psi, dtype=complex)
    psi, n0 = _normalize(psi)
    if n0 == 0:
        return None, {"fallback": True, "reason": "zero_input"}

    U = np.asarray(U, dtype=complex)
    backend = AerSimulator()
    Upows = _precompute_U_powers(U, n_bits)

    # bits theo thứ tự MSB -> LSB: [x1, x2, ..., xm]
    estimated_bits = [0] * n_bits
    rounds = []

    # lặp từ bit thấp nhất lên cao nhất theo kiểu semiclassical inverse QFT
    # vòng đầu dùng U^(2^(m-1)) để suy ra x_m
    for bit_index in range(n_bits - 1, -1, -1):
        U_power = Upows[bit_index]

        # tail bits đã biết: x_{bit_index+2} ... x_m
        tail_bits = estimated_bits[bit_index + 1:]
        feedback_frac = _binary_fraction_from_tail_bits(tail_bits)

        bit, round_info = _run_iqpe_round(
            backend=backend,
            U_power=U_power,
            psi=psi,
            feedback_frac=feedback_frac,
            shots=shots,
            optimization_level=optimization_level,
        )

        estimated_bits[bit_index] = bit

        rounds.append({
            "bit_position_msb1": bit_index + 1,
            "power": 2 ** bit_index,
            "feedback_frac": float(feedback_frac),
            "measured_bit": int(bit),
            **round_info,
        })

    phi_est = _bits_to_phase(estimated_bits)
    E_wrapped = _energy_from_phase(phi_est, t_evol)
    E_refined = _unwrap_to_reference(E_wrapped, E_target, t_evol)

    return float(E_refined), {
        "fallback": False,
        "bits_msb_first": estimated_bits,
        "bitstring_msb_first": "".join(str(int(b)) for b in estimated_bits),
        "phi_est": float(phi_est),
        "E_wrapped": float(E_wrapped),
        "E_refined": float(E_refined),
        "rounds": rounds,
    }


def _inverse_refine_vector(A, v0, lam, prev_vecs=None, n_iter=3, reg=1e-5):
    prev_vecs = prev_vecs or []
    A = np.asarray(A, dtype=float)
    v = np.asarray(np.real(v0), dtype=float)
    v = v / max(np.linalg.norm(v), 1e-12)

    I = np.eye(A.shape[0], dtype=float)

    for _ in range(n_iter):
        M = A - lam * I + reg * I
        x = np.linalg.solve(M, v)
        x = _deflate_against(x, prev_vecs)
        x = np.real(x)
        v = x / max(np.linalg.norm(x), 1e-12)

    lam_new = float(v @ A @ v)
    return lam_new, v.astype(complex)

# ============================================================
# MAIN PIPELINE for NCut: returns 3 sets of k eigenvectors
# ============================================================
def compute_ncut_ql_pipeline_three(
    A: np.ndarray,
    D_vals: np.ndarray,
    k: int,
    *,
    n_starts: int = 64,
    k_per_start: int = 8,
    m_krylov: int = 48,
    t_evol: float = 10.0,
    overlap_tol: float = 0.75,
    seed: int = 123,
    t_count: int = 10,
    topM: int = 16,
    w_energy: float = 1.0,
    w_overlap: float = 0.25,
    w_dup: float = 5.0,
    iqpe_powers: int = 10,
    iqpe_weights: str = "exp",
    iqpe_shots: int = 1024,
):
    start_V_ql = time.perf_counter()

    A = np.asarray(A, dtype=float)
    D_vals = np.asarray(D_vals, dtype=float)
    N = A.shape[0]

    n_qubits = int(np.log2(N))
    if 2**n_qubits != N:
        raise ValueError(f"N={N} must be power of 2 for Aer statevector QLanczos/QPE/IQPE.")

    # =========================================================
    # STEP 1: QL candidates
    # =========================================================
    n_candidates = max(3 * k + 6, k + 6)

    E_cand, V_cand = _ql_multistart_select(
        A,
        k=n_candidates,
        n_starts=n_starts,
        k_per_start=k_per_start,
        m_krylov=m_krylov,
        t_evol=t_evol,
        seed=seed,
        overlap_tol=overlap_tol,
    )
    if V_cand.shape[1] == 0:
        raise RuntimeError("Step1 QLanczos produced no candidates.")

    # =========================================================
    # STEP 2: QR -> Ritz refinement
    # =========================================================
    E_ritz, V_ritz = _ritz_refine_from_candidates(
        A=A,
        V_cand=V_cand,
        n_keep=min(V_cand.shape[1], max(2 * k + 4, k + 2)),
        qr_tol=1e-12,
    )

    if V_ritz.shape[1] == 0:
        raise RuntimeError("Ritz refinement returned no vectors.")

    # =========================================================
    # STEP 3: chọn non-trivial vectors cho normalized Laplacian
    # =========================================================
    E_ql, V_ql_c = _select_nontrivial_by_degree(
        E=E_ritz,
        V=V_ritz,
        D_vals=D_vals,
        k=k,
    )

    if V_ql_c.shape[1] == 0:
        raise RuntimeError("No non-trivial QL eigenvectors selected.")

    V_ql = _realify_columns(V_ql_c)

    U = expm(-1j * A * t_evol)

    end_V_ql = time.perf_counter()

    # =========================================================
    # STEP 4: QPE refine from refined QL vectors
    # =========================================================
    V_qpe_list = []
    E_qpe_list = []
    refined_vecs_for_dup = []

    k_eff = V_ql_c.shape[1]
    for i in range(k_eff):
        psi = V_ql_c[:, i]
        psi = _deflate_against(psi, refined_vecs_for_dup)
        psi_sv = Statevector(psi)

        v_ref, info = _qpe_postselect_best_mode(
            U=U,
            t_count=t_count,
            psi_in=psi_sv,
            E_target=float(E_ql[i]),
            t_evol=t_evol,
            prev_refined_vecs=refined_vecs_for_dup,
            topM=topM,
            w_energy=w_energy,
            w_overlap=w_overlap,
            w_dup=w_dup,
        )

        if v_ref is None:
            v_ref = psi
            E_qpe_list.append(float(E_ql[i]))
        else:
            E_qpe_list.append(float(info["E_refined"]))

        v_ref = _canonicalize_phase(v_ref)
        v_ref, _ = _normalize(v_ref)
        refined_vecs_for_dup.append(v_ref)
        V_qpe_list.append(v_ref)

    ######################## V_qpe_c = np.column_stack(V_qpe_list) if len(V_qpe_list) > 0 else np.zeros((N, 0), dtype=complex)
    ######################## E_qpe = np.array(E_qpe_list, dtype=float)
    ######################## V_qpe = _realify_columns(V_qpe_c)
    
    V_qpe_c = np.column_stack(V_qpe_list) if len(V_qpe_list) > 0 else np.zeros((N, 0), dtype=complex)
    # FINAL SUBSPACE REFINEMENT FOR QPE
    if V_qpe_c.shape[1] > 0:
        E_qpe_ref, V_qpe_ref = _ritz_refine_from_candidates(
            A=A,
            V_cand=V_qpe_c,
            n_keep=min(max(2 * k, k + 2), V_qpe_c.shape[1]),
            qr_tol=1e-12,
        )
    
        E_qpe, V_qpe_c = _select_nontrivial_by_degree(
            E=E_qpe_ref,
            V=V_qpe_ref,
            D_vals=D_vals,
            k=k,
        )
    else:
        E_qpe = np.array([], dtype=float)
        V_qpe_c = np.zeros((N, 0), dtype=complex)
    
    # ---------------------------------------------------------
    # FALLBACK: nếu QPE bị thiếu cột thì mượn thêm từ QL
    # ---------------------------------------------------------
    if V_qpe_c.shape[1] < k:
        missing = k - V_qpe_c.shape[1]
        extra_cols = []
        extra_E = []
    
        for j in range(V_ql_c.shape[1]):
            v = V_ql_c[:, j]
    
            if V_qpe_c.shape[1] > 0:
                dup = max(_overlap_abs(v, V_qpe_c[:, t]) for t in range(V_qpe_c.shape[1]))
            else:
                dup = 0.0
    
            if dup < 0.95:
                extra_cols.append(v.reshape(-1, 1))
                extra_E.append(float(E_ql[j]))
    
            if len(extra_cols) >= missing:
                break
    
        if len(extra_cols) > 0:
            if V_qpe_c.shape[1] > 0:
                V_qpe_c = np.column_stack([V_qpe_c] + extra_cols)
            else:
                V_qpe_c = np.column_stack(extra_cols)
    
            E_qpe = np.concatenate([E_qpe, np.array(extra_E[:missing], dtype=float)])
    
    E_qpe = E_qpe[:k]
    V_qpe_c = V_qpe_c[:, :k]
    
    idx = np.argsort(E_qpe)
    E_qpe = E_qpe[idx]
    V_qpe_c = V_qpe_c[:, idx]
    
    V_qpe = _realify_columns(V_qpe_c)
    
    end_V_qpe = time.perf_counter()

    # =========================================================
    # STEP 5: IQPE refine energy from refined QL vectors
    # iterative / semiclassical version with ancilla + measured rounds
    # =========================================================
    V_iqpe_list = []
    E_iqpe_list = []
    iqpe_selected_vecs = []

    for i in range(k_eff):
        psi = V_ql_c[:, i]
        psi = _deflate_against(psi, iqpe_selected_vecs)
        psi, _ = _normalize(psi)

        E_refined, _info = _iqpe_iterative_refine_energy(
            U=U,
            psi=psi,
            t_evol=t_evol,
            E_target=float(E_ql[i]),
            n_bits=iqpe_powers,          # giữ tên tham số cũ để khỏi sửa nhiều nơi
            shots=iqpe_shots,                  # có thể chỉnh 256 / 512 / 1024
            optimization_level=0,
        )

        if E_refined is None:
            E_iqpe_list.append(float(E_ql[i]))
        else:
            E_iqpe_list.append(float(E_refined))

        # vẫn giữ bước inverse refinement để suy vector từ energy estimate
        lam_used = float(E_refined) if E_refined is not None else float(E_ql[i])
        lam_new, v_new = _inverse_refine_vector(
            A, psi, lam_used,
            prev_vecs=iqpe_selected_vecs,
            n_iter=3,
            reg=1e-5
        )
        E_iqpe_list[-1] = lam_new

        v_new = _canonicalize_phase(v_new)
        v_new, _ = _normalize(v_new)
        iqpe_selected_vecs.append(v_new)
        V_iqpe_list.append(v_new)

    V_iqpe_c = np.column_stack(V_iqpe_list) if len(V_iqpe_list) > 0 else np.zeros((N, 0), dtype=complex)

    # FINAL SUBSPACE REFINEMENT FOR IQPE
    if V_iqpe_c.shape[1] > 0:
        E_iqpe_ref, V_iqpe_ref = _ritz_refine_from_candidates(
            A=A,
            V_cand=V_iqpe_c,
            n_keep=min(max(2 * k, k + 2), V_iqpe_c.shape[1]),
            qr_tol=1e-12,
        )
    
        E_iqpe, V_iqpe_c = _select_nontrivial_by_degree(
            E=E_iqpe_ref,
            V=V_iqpe_ref,
            D_vals=D_vals,
            k=k,
        )
    else:
        E_iqpe = np.array([], dtype=float)
        V_iqpe_c = np.zeros((N, 0), dtype=complex)
    
    # ---------------------------------------------------------
    # FALLBACK: nếu IQPE bị thiếu cột thì mượn thêm từ QL
    # ---------------------------------------------------------
    if V_iqpe_c.shape[1] < k:
        missing = k - V_iqpe_c.shape[1]
        extra_cols = []
        extra_E = []
    
        for j in range(V_ql_c.shape[1]):
            v = V_ql_c[:, j]
    
            if V_iqpe_c.shape[1] > 0:
                dup = max(_overlap_abs(v, V_iqpe_c[:, t]) for t in range(V_iqpe_c.shape[1]))
            else:
                dup = 0.0
    
            if dup < 0.95:
                extra_cols.append(v.reshape(-1, 1))
                extra_E.append(float(E_ql[j]))
    
            if len(extra_cols) >= missing:
                break
    
        if len(extra_cols) > 0:
            if V_iqpe_c.shape[1] > 0:
                V_iqpe_c = np.column_stack([V_iqpe_c] + extra_cols)
            else:
                V_iqpe_c = np.column_stack(extra_cols)
    
            E_iqpe = np.concatenate([E_iqpe, np.array(extra_E[:missing], dtype=float)])
    
    E_iqpe = E_iqpe[:k]
    V_iqpe_c = V_iqpe_c[:, :k]
    
    idx = np.argsort(E_iqpe)
    E_iqpe = E_iqpe[idx]
    V_iqpe_c = V_iqpe_c[:, idx]
    
    V_iqpe = _realify_columns(V_iqpe_c)

    end_V_iqpe = time.perf_counter()

    return E_ql, E_qpe, E_iqpe, V_ql, V_qpe, V_iqpe, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe

def compute_weight_matrix_coo_knn(image, sigma_i, sigma_x, k_neighbors=10):
    h, w, c = image.shape
    N = h * w

    coords = np.array(np.meshgrid(range(h), range(w))).reshape(2, -1).T
    features = image.reshape(-1, c)

    knn = NearestNeighbors(n_neighbors=k_neighbors, algorithm="ball_tree")
    knn.fit(coords)
    distances, indices = knn.kneighbors(coords)

    row_idx = np.repeat(np.arange(N), k_neighbors)
    col_idx = indices.flatten()

    feat_row = features[row_idx]
    feat_col = features[col_idx]

    diff = feat_row - feat_col
    W_feature = np.exp(-np.sum(diff ** 2, axis=1) / (2 * sigma_i ** 2))
    W_space = np.exp(-(distances.flatten() ** 2) / (2 * sigma_x ** 2))

    values = W_feature * W_space

    W_sparse = sp.coo_matrix((values, (row_idx, col_idx)), shape=(N, N))
    return W_sparse


def compute_laplacian_coo(W_coo):
    D = np.array(W_coo.sum(axis=1)).flatten()
    D_coo = sp.diags(D)
    L_coo = D_coo - W_coo
    return L_coo, D_coo

def assign_labels(eigen_vectors, k):
    return KMeans(n_clusters=k, random_state=0, n_init=20).fit(eigen_vectors).labels_

def save_segmentation(image, labels, k, output_path):
    h, w, c = image.shape
    segmented_image = np.zeros_like(image, dtype=np.uint8)
    for i in range(k):
        mask = labels.reshape(h, w) == i
        cluster_pixels = image[mask]
        mean_color = (cluster_pixels.mean(axis=0) * 255).astype(np.uint8) if len(cluster_pixels) > 0 else np.array([0, 0, 0], dtype=np.uint8)
        segmented_image[mask] = mean_color
    io.imsave(output_path, segmented_image)

def save_seg_file(labels, image_shape, output_path, image_name="image"):
    h, w = image_shape[:2]
    unique_labels = np.unique(labels)
    segments = len(unique_labels)

    # Tạo phần header
    header = [
        "format ascii cr",
        f"date {datetime.now().strftime('%a %b %d %H:%M:%S %Y')}",
        f"image ",
        "user 1102",  # Giữ nguyên như file mẫu
        f"width {w}",
        f"height {h}",
        f"segments {segments}",
        "gray 0",
        "invert 0",
        "flipflop 0",
        "data"
    ]

    # Tạo dữ liệu pixel theo định dạng (nhãn, dòng, cột bắt đầu, cột kết thúc)
    data_lines = []
    for row in range(h):
        row_labels = labels[row, :]
        start_col = 0
        current_label = row_labels[0]

        for col in range(1, w):
            if row_labels[col] != current_label:
                data_lines.append(f"{current_label} {row} {start_col} {col}")
                start_col = col
                current_label = row_labels[col]

        # Thêm dòng cuối cùng của hàng
        data_lines.append(f"{current_label} {row} {start_col} {w}")

    # Lưu vào file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(header) + "\n")
        f.write("\n".join(data_lines) + "\n")

    print(f"✅ File SEG đã lưu: {output_path}")

def align_eigenvector_signs(vecs, V_ql, V_qpe, V_iqpe):
    """
    Align signs of V_ql, V_qpe, V_iqpe to match reference eigenvectors vecs.
    Chỉ align trên số cột thực sự tồn tại của mỗi ma trận.
    """
    V_ql_aligned = V_ql.copy()
    V_qpe_aligned = V_qpe.copy()
    V_iqpe_aligned = V_iqpe.copy()

    k_ql = min(vecs.shape[1], V_ql_aligned.shape[1])
    k_qpe = min(vecs.shape[1], V_qpe_aligned.shape[1])
    k_iqpe = min(vecs.shape[1], V_iqpe_aligned.shape[1])

    for i in range(k_ql):
        ref = vecs[:, i]
        if np.dot(ref, V_ql_aligned[:, i]) < 0:
            V_ql_aligned[:, i] *= -1

    for i in range(k_qpe):
        ref = vecs[:, i]
        if np.dot(ref, V_qpe_aligned[:, i]) < 0:
            V_qpe_aligned[:, i] *= -1

    for i in range(k_iqpe):
        ref = vecs[:, i]
        if np.dot(ref, V_iqpe_aligned[:, i]) < 0:
            V_iqpe_aligned[:, i] *= -1

    return V_ql_aligned, V_qpe_aligned, V_iqpe_aligned

def build_ncut_matrix(W_coo):
    if not isinstance(W_coo, coo_matrix):
        W_coo = W_coo.tocoo()

    W_sym = (W_coo + W_coo.T).tocoo()
    W_sym.data *= 0.5

    D_vals = np.array(W_sym.sum(axis=1)).flatten()
    D_inv_sqrt = 1.0 / np.sqrt(D_vals + 1e-8)

    row, col = W_sym.row, W_sym.col
    data = W_sym.data * D_inv_sqrt[row] * D_inv_sqrt[col]
    W_norm = coo_matrix((data, (row, col)), shape=W_sym.shape).toarray()

    A = np.eye(W_coo.shape[0], dtype=float) - W_norm
    A = 0.5 * (A + A.T)
    return A, D_vals
    
def row_normalize(X, eps=1e-12):
    X = np.asarray(X, dtype=float)
    norms = np.linalg.norm(X, axis=1, keepdims=True)
    return X / np.maximum(norms, eps)
    
def normalized_cuts_eigsh(imagename, image_path, output_path, k, sigma_i, sigma_x):
    image = io.imread(image_path)
    image = color.gray2rgb(image) if image.ndim == 2 else image[:, :, :3] if image.shape[2] == 4 else image
    image = image / 255.0

    start_vecs = time.perf_counter()
    W_coo = compute_weight_matrix_coo_knn(image, sigma_i, sigma_x)

    A, D_vals = build_ncut_matrix(W_coo)
    A = np.asarray(A, dtype=float)

    # Traditional eigenpairs
    evals_all, vecs_all = np.linalg.eigh(A)
    idx_all = np.argsort(evals_all)
    evals_all = evals_all[idx_all]
    vecs_all = vecs_all[:, idx_all]

    evals, vecs_c = _select_nontrivial_by_degree(
        E=evals_all,
        V=vecs_all.astype(complex),
        D_vals=D_vals,
        k=k
    )
    vecs = _realify_columns(vecs_c)

    end_vecs = time.perf_counter()

    E_ql, E_qpe, E_iqpe, V_ql, V_qpe, V_iqpe, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe = compute_ncut_ql_pipeline_three(A, D_vals, k)

    V_ql, V_qpe, V_iqpe = align_eigenvector_signs(vecs, V_ql, V_qpe, V_iqpe)

    labels = assign_labels(row_normalize(vecs), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_L.seg", imagename)

    labels = assign_labels(row_normalize(V_ql), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_QL.seg", imagename)

    labels = assign_labels(row_normalize(V_qpe), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_QPE.seg", imagename)

    labels = assign_labels(row_normalize(V_iqpe), k)
    save_seg_file(labels.reshape(image.shape[:2]), image.shape, output_path + "_IQPE.seg", imagename)

    del W_coo
    return start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe, evals, E_ql, E_qpe, E_iqpe, vecs, V_ql, V_qpe, V_iqpe


def append_eigenvectors_row_format(
    excel_path,
    file_name,
    vecs,
    V_ql,
    V_qpe,
    V_iqpe,
    sheet_name="data"
):

    run_id = file_name   # run_id lấy trực tiếp từ file_name

    # ----- chuyển ma trận thành dataframe -----
    def pack(method, M):
        rows = []
        N, k = M.shape

        for j in range(k):
            row = {
                "run_id": run_id,
                "method": method,
                "eig_idx": j + 1
            }

            for i in range(N):
                row[f"v{i}"] = float(np.real(M[i, j]))

            rows.append(row)

        return rows

    rows = []
    rows += pack("Traditional", vecs)
    rows += pack("QLanczos", V_ql)
    rows += pack("QPE", V_qpe)
    rows += pack("IQPE", V_iqpe)

    df = pd.DataFrame(rows)

    # ----- nếu file chưa tồn tại → tạo mới -----
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ----- nếu file tồn tại → append -----
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )

    print("✅ Appended data to:", excel_path)

def append_eigenvalues_simple(
    excel_path,
    file_name,
    evals,
    E_ql,
    E_qpe,
    E_iqpe,
    sheet_name="data"
):

    run_id = file_name

    def pack(method, E):
        row = {
            "run_id": run_id,
            "method": method
        }

        for i, val in enumerate(E):
            row[f"e{i}"] = float(val)

        return row

    rows = [
        pack("Traditional", evals),
        pack("QLanczos", E_ql),
        pack("QPE", E_qpe),
        pack("IQPE", E_iqpe)
    ]

    df = pd.DataFrame(rows)

    # ---------- nếu file chưa tồn tại ----------
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ---------- nếu file tồn tại ----------
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,   # luôn append
            startrow=startrow
        )

    print("✅ Appended eigenvalues to:", excel_path)

def append_log_excel(
        excel_path,
        file_name,
        start_vecs,
        end_vecs,
        start_V_ql,
        end_V_ql,
        end_V_qpe,
        end_V_iqpe,
        sheet_name="log"
    ):

    # ----- tạo dataframe -----
    new_df = pd.DataFrame(
        [(file_name, start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe)],
        columns=[
            "Tên file",
            "start_L",
            "end_L",
            "start_QL",
            "end_QL",
            "end_QPE",
            "end_IQPE"
        ]
    )

    # ----- nếu file chưa tồn tại → tạo mới -----
    if not os.path.exists(excel_path):

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            new_df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✅ Created new Excel file:", excel_path)
        return

    # ----- file đã tồn tại -----
    book = load_workbook(excel_path)

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        startrow = sheet.max_row
    else:
        startrow = 0

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:

        new_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )

    print(f"📝 Appended data to: {excel_path}")

def main(name):
    numbers = re.findall(r'\d+', name)
    input_path = "/content/drive/MyDrive/TestK3/" + name
    excel_path = os.path.join("/content/drive/MyDrive/TestK3/log"+numbers[0]+".xlsx")  # file Excel lưu
    excel_path_Vector = os.path.join("/content/drive/MyDrive/TestK3/logVector"+numbers[0]+".xlsx")  # file Excel lưu
    excel_path_Eig = os.path.join("/content/drive/MyDrive/TestK3/logEig"+numbers[0]+".xlsx")  # file Excel lưu
    output_path = "/content/drive/MyDrive/TestK3/out"+numbers[0]

    if not os.path.isdir(input_path):
        print(f"❌ Thư mục {input_path} không tồn tại!")
        exit()
    os.makedirs(output_path, exist_ok=True)
    
    image_files = [f for f in os.listdir(input_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]
    if not image_files:
        print(f"❌ Không tìm thấy file ảnh nào trong {input_path}!")
        exit()

    log_rows = []  # mỗi phần tử: (tên file, bắt đầu, kết thúc)

    for idx, file_name in enumerate(image_files, start=1):
        start = time.perf_counter()
        k = int(re.search(r"_(\d+)\.png$", file_name).group(1))

        image_path = os.path.join(input_path, file_name)
        print(f"📷 Đang xử lý ảnh {idx}: {image_path}")

        sigma_i = 0.009
        sigma_x = 8

        save_image_name = os.path.join(output_path, f"{os.path.splitext(file_name)[0]}")
        start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe, evals, E_ql, E_qpe, E_iqpe, vecs, V_ql, V_qpe, V_iqpe = normalized_cuts_eigsh(file_name, image_path, save_image_name, k, sigma_i, sigma_x)
        
        append_log_excel(excel_path, file_name, start_vecs, end_vecs, start_V_ql, end_V_ql, end_V_qpe, end_V_iqpe)
        append_eigenvalues_simple(excel_path_Eig, file_name, evals, E_ql, E_qpe, E_iqpe)
        append_eigenvectors_row_format(excel_path_Vector, file_name, vecs, V_ql, V_qpe, V_iqpe)
        end = time.perf_counter()
        print("Thời gian xử lý 1 ảnh ",end-start)

#####################################      Chỗ sửa đường dẫn     ######################################################
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("nameImage", type=str)
    args = parser.parse_args()
    main(args.nameImage)
########################################################################################################################
